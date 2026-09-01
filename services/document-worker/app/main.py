from __future__ import annotations

import csv
import asyncio
import hashlib
import io
import json
import os
import re
import time
from typing import Annotated, Any

import fitz
import httpx
from fastapi import FastAPI, File, Form, Header, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

MAX_FILE_BYTES = int(os.getenv("DOCUMENT_MAX_BYTES", str(20 * 1024 * 1024)))
MAX_PDF_PAGES = int(os.getenv("DOCUMENT_MAX_PDF_PAGES", "80"))
MAX_TEXT_CHARS = int(os.getenv("DOCUMENT_MAX_TEXT_CHARS", "200000"))
MIN_NATIVE_TEXT_CHARS = int(os.getenv("DOCUMENT_MIN_NATIVE_TEXT_CHARS", "80"))
DOCLING_ENABLED = os.getenv("DOCUMENT_DOCLING_ENABLED", "true").lower() == "true"
DOCLING_API_URL = os.getenv("DOCLING_API_URL", "http://docling:5001").rstrip("/")
DOCLING_TIMEOUT_SECONDS = float(os.getenv("DOCLING_TIMEOUT_SECONDS", "300"))
DOCLING_POLL_INTERVAL_SECONDS = float(os.getenv("DOCLING_POLL_INTERVAL_SECONDS", "2"))
SUPPORTED_MIME_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

app = FastAPI(title="Diretor 360 Document Worker", version="1.2.0", docs_url=None, redoc_url=None)


class Evidence(BaseModel):
    model_config = ConfigDict(extra="forbid")
    locator: str
    text: str
    extraction_method: str
    confidence: float | None = Field(default=None, ge=0, le=1)


class ExtractedTable(BaseModel):
    model_config = ConfigDict(extra="forbid")
    table_id: str
    page_number: int | None = None
    headers: list[str]
    rows: list[list[str]]
    markdown: str
    locator: str
    warnings: list[str]


class ExtractedSection(BaseModel):
    model_config = ConfigDict(extra="forbid")
    level: int = Field(ge=1, le=6)
    title: str
    page_number: int | None = None
    locator: str


class ParserInfo(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name: str
    version: str | None = None
    processing_ms: int | None = Field(default=None, ge=0)


class Extraction(BaseModel):
    model_config = ConfigDict(extra="forbid")
    document_type: str
    mime_type: str
    file_name: str
    content_hash: str
    text: str
    markdown: str
    tables: list[ExtractedTable]
    sections: list[ExtractedSection]
    parser: ParserInfo
    page_count: int | None = None
    extraction_method: str
    evidence: list[Evidence]
    warnings: list[str]


class WorkerResponse(BaseModel):
    model_config = ConfigDict(extra="forbid")
    ok: bool
    schema_version: str = "1.1.0"
    worker_version: str = "1.2.0"
    job_id: str
    duration_ms: int
    security: dict[str, Any]
    extraction: Extraction


@app.get("/health")
def health() -> dict[str, str | bool]:
    return {
        "status": "ok",
        "service": "document-worker",
        "version": "1.2.0",
        "docling_enabled": DOCLING_ENABLED,
    }


@app.post("/v1/process", response_model=WorkerResponse)
async def process_document(
    document: Annotated[UploadFile, File()],
    metadata: Annotated[str, Form()],
    x_job_id: Annotated[str | None, Header()] = None,
    x_content_trust: Annotated[str | None, Header()] = None,
) -> WorkerResponse:
    started = time.monotonic()
    job = parse_metadata(metadata)
    job_id = x_job_id or str(job.get("job_id", ""))
    if not job_id or len(job_id) > 160:
        raise HTTPException(400, "invalid_job_id")
    if x_content_trust != "UNTRUSTED" or job.get("security", {}).get("external_effects_allowed") is not False:
        raise HTTPException(400, "invalid_security_context")

    mime = (document.content_type or "").lower()
    if mime not in SUPPORTED_MIME_TYPES:
        raise HTTPException(415, "unsupported_media_type")
    content = await read_bounded(document)
    validate_signature(content, mime)
    digest = f"sha256:{hashlib.sha256(content).hexdigest()}"
    expected_hash = str(job.get("document", {}).get("content_hash") or "")
    if expected_hash and expected_hash != digest:
        raise HTTPException(422, "content_hash_mismatch")

    file_name = safe_name(document.filename or job.get("document", {}).get("file_name") or "documento")
    if mime == "application/pdf":
        extracted = await extract_pdf_routed(content, file_name, digest)
    elif mime in {"image/jpeg", "image/png"}:
        extracted = await extract_image_routed(content, file_name, mime, digest)
    elif mime == "text/csv":
        extracted = extract_csv(content, file_name, digest)
    else:
        extracted = extract_xlsx(content, file_name, digest)

    duration_ms = round((time.monotonic() - started) * 1000)
    return WorkerResponse(
        ok=True,
        job_id=job_id,
        duration_ms=duration_ms,
        security={
            "content_trust": "UNTRUSTED",
            "instructions_in_document_ignored": True,
            "external_effects_allowed": False,
        },
        extraction=extracted,
    )


async def extract_pdf_routed(content: bytes, file_name: str, digest: str) -> Extraction:
    if DOCLING_ENABLED:
        try:
            return await extract_with_docling(content, file_name, "application/pdf", digest)
        except (httpx.HTTPError, KeyError, TypeError, ValueError, json.JSONDecodeError, TimeoutError) as error:
            fallback = extract_pdf(content, file_name, digest)
            fallback.warnings.insert(0, f"docling_fallback:{type(error).__name__}")
            if fallback.evidence:
                return fallback
            raise HTTPException(503, "docling_required_for_scanned_pdf") from error
    return extract_pdf(content, file_name, digest)


async def extract_image_routed(content: bytes, file_name: str, mime: str, digest: str) -> Extraction:
    if not DOCLING_ENABLED:
        raise HTTPException(503, "docling_required_for_image")
    try:
        return await extract_with_docling(content, file_name, mime, digest)
    except (httpx.HTTPError, KeyError, TypeError, ValueError, json.JSONDecodeError, TimeoutError) as error:
        raise HTTPException(503, "docling_image_processing_failed") from error


async def extract_with_docling(content: bytes, file_name: str, mime: str, digest: str) -> Extraction:
    started = time.monotonic()
    request_data = {
        "to_formats": ["md", "json"],
        "image_export_mode": "placeholder", "do_ocr": "true",
        "force_ocr": "false", "ocr_lang": ["pt", "en"],
        "do_table_structure": "true", "table_mode": "accurate",
        "table_cell_matching": "true", "include_images": "false",
        "include_page_images": "false", "abort_on_error": "false",
        "document_timeout": str(int(DOCLING_TIMEOUT_SECONDS)),
    }
    timeout = httpx.Timeout(DOCLING_TIMEOUT_SECONDS + 15, connect=10.0)
    async with httpx.AsyncClient(timeout=timeout) as client:
        response = await client.post(
            f"{DOCLING_API_URL}/v1/convert/file/async",
            data=request_data,
            files={"files": (file_name, content, mime)},
        )
        response.raise_for_status()
        task_id = str(response.json().get("task_id") or "")
        if not task_id:
            raise ValueError("docling_task_missing")
        deadline = time.monotonic() + DOCLING_TIMEOUT_SECONDS
        while time.monotonic() < deadline:
            status_response = await client.get(f"{DOCLING_API_URL}/v1/status/poll/{task_id}")
            status_response.raise_for_status()
            status = str(status_response.json().get("task_status") or status_response.json().get("status") or "").lower()
            if status in {"success", "partial_success", "failure", "failed"}:
                break
            await asyncio.sleep(DOCLING_POLL_INTERVAL_SECONDS)
        else:
            raise TimeoutError("docling_timeout")
        if status in {"failure", "failed"}:
            raise ValueError("docling_processing_failed")
        result_response = await client.get(f"{DOCLING_API_URL}/v1/result/{task_id}")
        result_response.raise_for_status()
        payload = result_response.json()
    return parse_docling_result(payload, file_name, mime, digest, status, round((time.monotonic() - started) * 1000))


def parse_docling_result(payload: dict[str, Any], file_name: str, mime: str, digest: str, status: str, duration_ms: int) -> Extraction:
    document = payload.get("document") or payload.get("result", {}).get("document") or {}
    markdown_warnings: list[str] = []
    markdown = bounded_text(str(document.get("md_content") or ""), markdown_warnings)
    doc_json = document.get("json_content") or {}
    if isinstance(doc_json, str):
        doc_json = json.loads(doc_json)
    if not isinstance(doc_json, dict):
        raise TypeError("docling_json_invalid")
    tables = parse_docling_tables(doc_json)
    sections = parse_docling_sections(doc_json)
    warnings = list(markdown_warnings)
    if status == "partial_success":
        warnings.append("docling_partial_success")
    for table in tables:
        warnings.extend(item for item in table.warnings if item not in warnings)
    if not markdown:
        warnings.append("docling_without_useful_text")
    evidence = [Evidence(locator=table.locator, text=table.markdown, extraction_method="DOCLING_TABLEFORMER") for table in tables]
    if not evidence and markdown:
        evidence.append(Evidence(locator="document:markdown", text=markdown, extraction_method="DOCLING_OCR"))
    page_numbers = [value for value in [table.page_number for table in tables] + [section.page_number for section in sections] if value]
    page_count = max(page_numbers) if page_numbers else (1 if mime.startswith("image/") else None)
    method = "DOCLING_TABLEFORMER" if tables else "DOCLING_OCR"
    version = str(payload.get("service_version") or payload.get("version") or "1.30.0")
    return Extraction(document_type="PDF" if mime == "application/pdf" else "IMAGE", mime_type=mime,
        file_name=file_name, content_hash=digest, text=markdown, markdown=markdown, tables=tables,
        sections=sections, parser=ParserInfo(name="DOCLING", version=version, processing_ms=duration_ms),
        page_count=page_count, extraction_method=method, evidence=evidence, warnings=warnings)


def parse_docling_tables(doc_json: dict[str, Any]) -> list[ExtractedTable]:
    parsed: list[ExtractedTable] = []
    headers_by_width: dict[int, list[str]] = {}
    for index, raw in enumerate(doc_json.get("tables") or [], start=1):
        if not isinstance(raw, dict):
            continue
        data = raw.get("data") or {}
        cells = data.get("table_cells") or raw.get("table_cells") or []
        row_count = int(data.get("num_rows") or raw.get("num_rows") or 0)
        col_count = int(data.get("num_cols") or raw.get("num_cols") or 0)
        warnings: list[str] = []
        logical_cells: list[tuple[int, int, int, int, str, bool]] = []
        for cell in cells:
            if not isinstance(cell, dict):
                continue
            rs, re_ = int(cell.get("start_row_offset_idx", 0)), int(cell.get("end_row_offset_idx", 0))
            cs, ce = int(cell.get("start_col_offset_idx", 0)), int(cell.get("end_col_offset_idx", 0))
            if re_ - rs > 1 or ce - cs > 1:
                warnings.append("docling_merged_cells")
            text = str(cell.get("text") or "").strip()
            logical_cells.append((rs, re_, cs, ce, text, bool(cell.get("column_header"))))
        if logical_cells:
            row_count = max(row_count, max(cell[1] for cell in logical_cells))
            col_count = max(col_count, max(cell[3] for cell in logical_cells))
        # Preserve physical offsets. Removing empty cells silently shifts the
        # business columns whenever Docling reports a merge or an empty cell.
        # Merged content stays at its first column and covered columns remain
        # blank, so downstream validation can ask Rafael instead of guessing.
        grid = [["" for _ in range(col_count)] for _ in range(row_count)]
        for rs, _re, cs, _ce, text, _is_header in logical_cells:
            if rs >= row_count or cs >= col_count:
                warnings.append("docling_incomplete_table")
                continue
            if grid[rs][cs] and text and grid[rs][cs] != text:
                warnings.append("docling_overlapping_cells")
                continue
            grid[rs][cs] = text
        compact_rows = grid
        if not compact_rows:
            warnings.append("docling_incomplete_table")
        header_index = next((row_index for row_index, row in enumerate(compact_rows)
            if sum(1 for value in row if normalize_header(value) in {"produto", "indicador", "peso", "metrica", "meta", "realizado", "%ating", "pontos"}) >= 3), None)
        headers = compact_rows[header_index] if header_index is not None else []
        rows = compact_rows[header_index + 1:] if header_index is not None else compact_rows
        common_width = col_count or max((len(row) for row in rows), default=len(headers))
        if headers:
            headers_by_width[common_width] = headers
        elif common_width in headers_by_width:
            headers = headers_by_width[common_width]
        rows = [row for row in rows if sum(1 for value in row if normalize_header(value) in {"produto", "indicador", "peso", "metrica", "meta", "realizado", "%ating", "pontos"}) < 3]
        if rows and any(abs(len(row) - common_width) > 1 for row in rows):
            warnings.append("docling_incomplete_table")
        normalized_headers = {normalize_header(value) for value in headers}
        normalized_header_list = [normalize_header(value) for value in headers]
        date_index = next((position for position, value in enumerate(normalized_header_list) if value in {"dtbase", "database"}), None)
        metric_index = next((position for position, value in enumerate(normalized_header_list) if value == "metrica"), None)
        for row in rows:
            if headers and headers[0].strip() == "+" and len(row) > 1 and row[0] and not row[1]:
                warnings.append("docling_possible_column_shift")
            if date_index is not None and date_index > 0 and len(row) > date_index:
                if not row[date_index] and re.fullmatch(r"\d{2}/\d{2}/\d{4}", row[date_index - 1].strip()):
                    warnings.append("docling_possible_column_shift")
            if metric_index is not None and len(row) > metric_index:
                if re.fullmatch(r"\d{2}/\d{2}/\d{4}", row[metric_index].strip()):
                    warnings.append("docling_possible_column_shift")
        critical = {"meta", "realizado"}
        pobj_like = any(len(row) >= 9 and any(normalize_header(value) in {"qtd", "prod", "perc", "pont"} for value in row) for row in rows)
        if (normalized_headers & critical and not critical.issubset(normalized_headers)) or (pobj_like and not headers):
            warnings.append("docling_possible_column_shift")
        provenance = raw.get("prov") or []
        page = next((int(item.get("page_no")) for item in provenance if isinstance(item, dict) and item.get("page_no")), None)
        locator = f"page:{page};table:{index}" if page else f"table:{index}"
        display_headers = headers or [f"col_{position + 1}" for position in range(common_width)]
        markdown = table_to_markdown(display_headers, rows)
        parsed.append(ExtractedTable(table_id=f"table-{index}", page_number=page, headers=headers, rows=rows,
            markdown=markdown, locator=locator, warnings=list(dict.fromkeys(warnings))))
    return parsed


def parse_docling_sections(doc_json: dict[str, Any]) -> list[ExtractedSection]:
    sections: list[ExtractedSection] = []
    for index, item in enumerate(doc_json.get("texts") or [], start=1):
        if not isinstance(item, dict) or item.get("label") not in {"title", "section_header"}:
            continue
        title = str(item.get("text") or "").strip()
        if not title:
            continue
        provenance = item.get("prov") or []
        page = next((int(value.get("page_no")) for value in provenance if isinstance(value, dict) and value.get("page_no")), None)
        level = 1 if item.get("label") == "title" else 2
        sections.append(ExtractedSection(level=level, title=title, page_number=page, locator=f"page:{page};section:{index}" if page else f"section:{index}"))
    return sections


def normalize_header(value: str) -> str:
    import unicodedata
    return re.sub(r"[^a-z0-9%]+", "", unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower())


def table_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    if not headers:
        return ""
    escape = lambda value: str(value).replace("|", "\\|").replace("\n", " ")
    lines = ["| " + " | ".join(map(escape, headers)) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    lines.extend("| " + " | ".join(map(escape, row)) + " |" for row in rows)
    return "\n".join(lines)


def parse_metadata(value: str) -> dict[str, Any]:
    try:
        parsed = json.loads(value)
    except (json.JSONDecodeError, TypeError) as error:
        raise HTTPException(400, "invalid_metadata") from error
    if not isinstance(parsed, dict):
        raise HTTPException(400, "invalid_metadata")
    return parsed


async def read_bounded(upload: UploadFile) -> bytes:
    content = await upload.read(MAX_FILE_BYTES + 1)
    if not content or len(content) > MAX_FILE_BYTES:
        raise HTTPException(413, "invalid_file_size")
    return content


def validate_signature(content: bytes, mime: str) -> None:
    valid = {
        "application/pdf": content.startswith(b"%PDF-"),
        "image/jpeg": content.startswith(b"\xff\xd8\xff"),
        "image/png": content.startswith(b"\x89PNG\r\n\x1a\n"),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": content.startswith(b"PK\x03\x04"),
        "text/csv": b"\x00" not in content[:8192],
    }.get(mime, False)
    if not valid:
        raise HTTPException(422, "file_content_mismatch")


def extract_pdf(content: bytes, file_name: str, digest: str) -> Extraction:
    document = fitz.open(stream=content, filetype="pdf")
    if document.page_count > MAX_PDF_PAGES:
        raise HTTPException(422, "pdf_page_limit_exceeded")
    evidence: list[Evidence] = []
    warnings: list[str] = []
    for page_number, page in enumerate(document, start=1):
        native_text = normalize_text(page.get_text("text"))
        if len(native_text) >= MIN_NATIVE_TEXT_CHARS:
            text = native_text
            method = "PDF_NATIVE_TEXT"
            confidence = None
        else:
            text = ""
            method = "PDF_NATIVE_TEXT"
            confidence = None
            warnings.append(f"page_{page_number}_requires_docling_ocr")
        if text:
            evidence.append(Evidence(locator=f"page:{page_number}", text=text, extraction_method=method, confidence=confidence))
    full_text = bounded_text("\n\n".join(item.text for item in evidence), warnings)
    return Extraction(
        document_type="PDF", mime_type="application/pdf", file_name=file_name, content_hash=digest,
        text=full_text, markdown=full_text, tables=[], sections=[],
        parser=ParserInfo(name="NATIVE", version=fitz.VersionBind, processing_ms=None), page_count=document.page_count,
        extraction_method="PDF_NATIVE_TEXT" if evidence else "NO_TEXT",
        evidence=evidence, warnings=warnings,
    )


def extract_csv(content: bytes, file_name: str, digest: str) -> Extraction:
    decoded = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(decoded)))[:5000]
    lines = [" | ".join(cell.strip() for cell in row) for row in rows if any(cell.strip() for cell in row)]
    text = bounded_text("\n".join(lines), [])
    evidence = [Evidence(locator=f"row:{index}", text=line, extraction_method="CSV_NATIVE") for index, line in enumerate(lines, start=1)]
    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    table = ExtractedTable(table_id="table-1", page_number=None, headers=headers, rows=data_rows,
                           markdown=table_to_markdown(headers, data_rows), locator="csv:table:1", warnings=[])
    return Extraction(document_type="CSV", mime_type="text/csv", file_name=file_name, content_hash=digest, text=text,
                      markdown=table.markdown, tables=[table] if headers else [], sections=[],
                      parser=ParserInfo(name="NATIVE", version=None, processing_ms=None),
                      extraction_method="CSV_NATIVE", evidence=evidence[:1000], warnings=[])


def extract_xlsx(content: bytes, file_name: str, digest: str) -> Extraction:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    evidence: list[Evidence] = []
    lines: list[str] = []
    tables: list[ExtractedTable] = []
    for sheet in workbook.worksheets[:20]:
        sheet_rows: list[list[str]] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value).strip() for value in row]
            if not any(values):
                continue
            sheet_rows.append(values)
            line = " | ".join(values)
            lines.append(line)
            if len(evidence) < 5000:
                evidence.append(Evidence(locator=f"sheet:{sheet.title};row:{row_index}", text=line, extraction_method="XLSX_NATIVE"))
            if len(lines) >= 10000:
                break
        if sheet_rows:
            headers, data_rows = sheet_rows[0], sheet_rows[1:]
            tables.append(ExtractedTable(table_id=f"sheet-{len(tables) + 1}", page_number=None,
                headers=headers, rows=data_rows, markdown=table_to_markdown(headers, data_rows),
                locator=f"sheet:{sheet.title}", warnings=[]))
    warnings: list[str] = []
    text = bounded_text("\n".join(lines), warnings)
    return Extraction(document_type="XLSX", mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      file_name=file_name, content_hash=digest, text=text, markdown="\n\n".join(table.markdown for table in tables),
                      tables=tables, sections=[], parser=ParserInfo(name="NATIVE", version=None, processing_ms=None), extraction_method="XLSX_NATIVE",
                      evidence=evidence, warnings=warnings)


def bounded_text(value: str, warnings: list[str]) -> str:
    if len(value) <= MAX_TEXT_CHARS:
        return value
    warnings.append("text_truncated")
    return value[:MAX_TEXT_CHARS]


def normalize_text(value: str) -> str:
    return "\n".join(" ".join(line.split()) for line in value.splitlines() if line.strip())


def safe_name(value: str) -> str:
    return "".join("_" if char in "\\/\x00" or ord(char) < 32 else char for char in value)[:120]
